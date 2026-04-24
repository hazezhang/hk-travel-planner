import glob
import hashlib
import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).parent
DB_PATH = ROOT / "food_rag.db"


def _find_food_excel_path() -> Path:
    matches = sorted(glob.glob(str(ROOT / "*.xlsx")))
    for p in matches:
        name = Path(p).name
        if "美食" in name or "food" in name.lower():
            return Path(p)
    raise FileNotFoundError("Food .xlsx source not found (expected name containing 美食/food).")


def _find_poi_excel_path() -> Path:
    matches = sorted(glob.glob(str(ROOT / "*.xlsx")))
    for p in matches:
        name = Path(p).name
        if "酒店+景点" in name or ("酒店" in name and "景点" in name):
            return Path(p)
    raise FileNotFoundError("POI .xlsx source not found (expected name containing 酒店+景点).")


def _file_md5(path: Path) -> str:
    h = hashlib.md5()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _ensure_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS food_places (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cuisine TEXT,
            restaurant TEXT,
            location TEXT,
            dishes_flavor TEXT,
            price TEXT,
            ambience_hygiene TEXT,
            service TEXT,
            value_note TEXT,
            flavor_by_cuisine TEXT,
            searchable_text TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS poi_places (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            attraction TEXT,
            location TEXT,
            category TEXT,
            highlights TEXT,
            duration TEXT,
            price TEXT,
            suitable_for TEXT,
            experience_tags TEXT,
            searchable_text TEXT
        )
        """
    )
    conn.commit()


def _rebuild_from_excel(conn: sqlite3.Connection, excel_path: Path) -> int:
    wb = load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Food Excel has no data rows.")

    conn.execute("DELETE FROM food_places")
    inserted = 0
    for row in rows[1:]:
        cells = [(str(v).strip() if v is not None else "") for v in row]
        if len(cells) < 9:
            cells += [""] * (9 - len(cells))
        cuisine, restaurant, location, dishes_flavor, price, ambience_hygiene, service, value_note, flavor_by_cuisine = cells[:9]
        if not restaurant:
            continue
        searchable = " ".join(
            [
                cuisine,
                restaurant,
                location,
                dishes_flavor,
                price,
                ambience_hygiene,
                service,
                value_note,
                flavor_by_cuisine,
            ]
        ).lower()
        conn.execute(
            """
            INSERT INTO food_places (
                cuisine, restaurant, location, dishes_flavor, price,
                ambience_hygiene, service, value_note, flavor_by_cuisine, searchable_text
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                cuisine,
                restaurant,
                location,
                dishes_flavor,
                price,
                ambience_hygiene,
                service,
                value_note,
                flavor_by_cuisine,
                searchable,
            ),
        )
        inserted += 1

    conn.execute(
        "INSERT OR REPLACE INTO meta(key, value) VALUES('food_source_file', ?)",
        (excel_path.name,),
    )
    conn.execute(
        "INSERT OR REPLACE INTO meta(key, value) VALUES('food_source_md5', ?)",
        (_file_md5(excel_path),),
    )
    conn.commit()
    return inserted


def _rebuild_poi_from_excel(conn: sqlite3.Connection, excel_path: Path) -> int:
    wb = load_workbook(excel_path)
    ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("POI Excel has no data rows.")

    conn.execute("DELETE FROM poi_places")
    inserted = 0
    for row in rows[1:]:
        cells = [(str(v).strip() if v is not None else "") for v in row]
        if len(cells) < 8:
            cells += [""] * (8 - len(cells))
        attraction, location, category, highlights, duration, price, suitable_for, experience_tags = cells[:8]
        if not attraction:
            continue
        searchable = " ".join(
            [attraction, location, category, highlights, duration, price, suitable_for, experience_tags]
        ).lower()
        conn.execute(
            """
            INSERT INTO poi_places (
                attraction, location, category, highlights, duration, price,
                suitable_for, experience_tags, searchable_text
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                attraction,
                location,
                category,
                highlights,
                duration,
                price,
                suitable_for,
                experience_tags,
                searchable,
            ),
        )
        inserted += 1

    conn.execute(
        "INSERT OR REPLACE INTO meta(key, value) VALUES('poi_source_file', ?)",
        (excel_path.name,),
    )
    conn.execute(
        "INSERT OR REPLACE INTO meta(key, value) VALUES('poi_source_md5', ?)",
        (_file_md5(excel_path),),
    )
    conn.commit()
    return inserted


def _rebuild_hotel_from_excel(conn: sqlite3.Connection, excel_path: Path) -> int:
    wb = load_workbook(excel_path)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Hotel Excel has no data rows.")

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS hotel_places (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            hotel TEXT,
            location TEXT,
            star_position TEXT,
            highlights TEXT,
            price_range TEXT,
            suitable_for TEXT,
            nearby TEXT,
            experience_tags TEXT,
            searchable_text TEXT
        )
        """
    )
    conn.execute("DELETE FROM hotel_places")
    inserted = 0
    for row in rows[1:]:
        cells = [(str(v).strip() if v is not None else "") for v in row]
        if len(cells) < 8:
            cells += [""] * (8 - len(cells))
        hotel, location, star_position, highlights, price_range, suitable_for, nearby, experience_tags = cells[:8]
        if not hotel:
            continue
        searchable = " ".join(
            [hotel, location, star_position, highlights, price_range, suitable_for, nearby, experience_tags]
        ).lower()
        conn.execute(
            """
            INSERT INTO hotel_places (
                hotel, location, star_position, highlights, price_range,
                suitable_for, nearby, experience_tags, searchable_text
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                hotel,
                location,
                star_position,
                highlights,
                price_range,
                suitable_for,
                nearby,
                experience_tags,
                searchable,
            ),
        )
        inserted += 1

    conn.execute(
        "INSERT OR REPLACE INTO meta(key, value) VALUES('hotel_source_file', ?)",
        (excel_path.name,),
    )
    conn.execute(
        "INSERT OR REPLACE INTO meta(key, value) VALUES('hotel_source_md5', ?)",
        (_file_md5(excel_path),),
    )
    conn.commit()
    return inserted


def ensure_food_database() -> dict[str, Any]:
    excel_path = _find_food_excel_path()
    excel_md5 = _file_md5(excel_path)
    with sqlite3.connect(DB_PATH) as conn:
        _ensure_schema(conn)
        old_md5_row = conn.execute("SELECT value FROM meta WHERE key='food_source_md5'").fetchone()
        old_md5 = old_md5_row[0] if old_md5_row else ""
        if old_md5 != excel_md5:
            count = _rebuild_from_excel(conn, excel_path)
            return {"status": "rebuilt", "rows": count, "source": excel_path.name}
        count = conn.execute("SELECT COUNT(*) FROM food_places").fetchone()[0]
        return {"status": "ready", "rows": count, "source": excel_path.name}


def ensure_poi_database() -> dict[str, Any]:
    excel_path = _find_poi_excel_path()
    excel_md5 = _file_md5(excel_path)
    with sqlite3.connect(DB_PATH) as conn:
        _ensure_schema(conn)
        old_md5_row = conn.execute("SELECT value FROM meta WHERE key='poi_source_md5'").fetchone()
        old_md5 = old_md5_row[0] if old_md5_row else ""
        if old_md5 != excel_md5:
            count = _rebuild_poi_from_excel(conn, excel_path)
            return {"status": "rebuilt", "rows": count, "source": excel_path.name}
        count = conn.execute("SELECT COUNT(*) FROM poi_places").fetchone()[0]
        return {"status": "ready", "rows": count, "source": excel_path.name}


def ensure_hotel_database() -> dict[str, Any]:
    excel_path = _find_poi_excel_path()
    excel_md5 = _file_md5(excel_path)
    with sqlite3.connect(DB_PATH) as conn:
        _ensure_schema(conn)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS hotel_places (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                hotel TEXT,
                location TEXT,
                star_position TEXT,
                highlights TEXT,
                price_range TEXT,
                suitable_for TEXT,
                nearby TEXT,
                experience_tags TEXT,
                searchable_text TEXT
            )
            """
        )
        old_md5_row = conn.execute("SELECT value FROM meta WHERE key='hotel_source_md5'").fetchone()
        old_md5 = old_md5_row[0] if old_md5_row else ""
        if old_md5 != excel_md5:
            count = _rebuild_hotel_from_excel(conn, excel_path)
            return {"status": "rebuilt", "rows": count, "source": excel_path.name}
        count = conn.execute("SELECT COUNT(*) FROM hotel_places").fetchone()[0]
        return {"status": "ready", "rows": count, "source": excel_path.name}


def search_food_places(
    query: str = "",
    budget_level: str = "",
    preferred_area: str = "",
    preferred_cuisines: list[str] | None = None,
    top_k: int = 8,
) -> dict[str, Any]:
    ensure_food_database()
    preferred_cuisines = preferred_cuisines or []
    q_tokens = [t.strip().lower() for t in query.split() if t.strip()]
    area_lc = preferred_area.lower().strip()
    cuisines_lc = [c.lower().strip() for c in preferred_cuisines if c.strip()]

    budget_hint = {
        "low": ["low", "budget", "cheap", "affordable", "��", "����", "street", "local"],
        "medium": ["medium", "moderate", "casual", "mid"],
        "high": ["high", "fine", "premium", "luxury", "expensive", "�ϸ�"],
    }.get((budget_level or "").lower(), [])

    with sqlite3.connect(DB_PATH) as conn:
        rows = conn.execute(
            """
            SELECT cuisine, restaurant, location, dishes_flavor, price,
                   ambience_hygiene, service, value_note, flavor_by_cuisine, searchable_text
            FROM food_places
            """
        ).fetchall()

    scored = []
    for r in rows:
        cuisine, restaurant, location, dishes_flavor, price, ambience_hygiene, service, value_note, flavor_by_cuisine, searchable = r
        score = 0
        text = searchable or ""

        for tok in q_tokens:
            if tok in text:
                score += 2
        if area_lc and area_lc in (location or "").lower():
            score += 3
        for c in cuisines_lc:
            if c in (cuisine or "").lower():
                score += 3
        for b in budget_hint:
            if b in text:
                score += 1
        if "local_food" in q_tokens and ("local" in text or "street" in text):
            score += 2

        if score > 0 or not q_tokens:
            scored.append(
                {
                    "restaurant": restaurant,
                    "cuisine": cuisine,
                    "location": location,
                    "price": price,
                    "dishes_flavor": dishes_flavor,
                    "ambience_hygiene": ambience_hygiene,
                    "service": service,
                    "value_note": value_note,
                    "flavor_by_cuisine": flavor_by_cuisine,
                    "_score": score,
                }
            )

    scored.sort(key=lambda x: x["_score"], reverse=True)
    top = scored[: max(1, min(top_k, 20))]
    for item in top:
        item.pop("_score", None)
    return {
        "source": "xlsx->sqlite_rag",
        "total_candidates": len(scored),
        "results": top,
    }


def search_poi_places(
    query: str = "",
    preferred_area: str = "",
    preferred_categories: list[str] | None = None,
    top_k: int = 10,
) -> dict[str, Any]:
    ensure_poi_database()
    preferred_categories = preferred_categories or []
    q_tokens = [t.strip().lower() for t in query.split() if t.strip()]
    area_lc = preferred_area.lower().strip()
    cat_lc = [c.lower().strip() for c in preferred_categories if c.strip()]

    with sqlite3.connect(DB_PATH) as conn:
        rows = conn.execute(
            """
            SELECT attraction, location, category, highlights, duration, price,
                   suitable_for, experience_tags, searchable_text
            FROM poi_places
            """
        ).fetchall()

    scored = []
    for r in rows:
        attraction, location, category, highlights, duration, price, suitable_for, experience_tags, searchable = r
        score = 0
        text = searchable or ""
        for tok in q_tokens:
            if tok in text:
                score += 2
        if area_lc and area_lc in (location or "").lower():
            score += 3
        for c in cat_lc:
            if c in (category or "").lower():
                score += 3
        if score > 0 or not q_tokens:
            scored.append(
                {
                    "attraction": attraction,
                    "location": location,
                    "category": category,
                    "highlights": highlights,
                    "duration": duration,
                    "price": price,
                    "suitable_for": suitable_for,
                    "experience_tags": experience_tags,
                    "_score": score,
                }
            )

    scored.sort(key=lambda x: x["_score"], reverse=True)
    top = scored[: max(1, min(top_k, 20))]
    for item in top:
        item.pop("_score", None)
    return {
        "source": "xlsx->sqlite_rag",
        "total_candidates": len(scored),
        "results": top,
    }


def search_hotel_places(
    query: str = "",
    budget_level: str = "",
    preferred_area: str = "",
    top_k: int = 6,
) -> dict[str, Any]:
    ensure_hotel_database()
    q_tokens = [t.strip().lower() for t in query.split() if t.strip()]
    area_lc = preferred_area.lower().strip()
    budget_hint = {
        "low": ["budget", "cheap", "affordable", "$", "economy", "hostel"],
        "medium": ["mid", "moderate", "4-star", "value"],
        "high": ["luxury", "5-star", "premium", "high", "$$$"],
    }.get((budget_level or "").lower(), [])

    with sqlite3.connect(DB_PATH) as conn:
        rows = conn.execute(
            """
            SELECT hotel, location, star_position, highlights, price_range,
                   suitable_for, nearby, experience_tags, searchable_text
            FROM hotel_places
            """
        ).fetchall()

    scored = []
    for r in rows:
        hotel, location, star_position, highlights, price_range, suitable_for, nearby, experience_tags, searchable = r
        score = 0
        text = searchable or ""
        for tok in q_tokens:
            if tok in text:
                score += 2
        if area_lc and area_lc in (location or "").lower():
            score += 3
        for b in budget_hint:
            if b in text:
                score += 1
        if score > 0 or not q_tokens:
            scored.append(
                {
                    "hotel": hotel,
                    "location": location,
                    "star_position": star_position,
                    "highlights": highlights,
                    "price_range": price_range,
                    "suitable_for": suitable_for,
                    "nearby": nearby,
                    "experience_tags": experience_tags,
                    "_score": score,
                }
            )

    scored.sort(key=lambda x: x["_score"], reverse=True)
    top = scored[: max(1, min(top_k, 20))]
    for item in top:
        item.pop("_score", None)
    return {
        "source": "xlsx->sqlite_rag",
        "total_candidates": len(scored),
        "results": top,
    }

